from io import BytesIO
from pathlib import Path
import tempfile
import pandas as pd
import plotly.express as px
import streamlit as st

from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import Output_Simple as planner


# =========================
# Page config
# =========================
APP_DIR = Path(__file__).parent
SAMPLE_INPUT = APP_DIR / "input_sample.xlsx"
LOCAL_INPUT = APP_DIR / "All_Output1.xlsx"

st.set_page_config(
    page_title="Output Simple Planner",
    layout="wide",
    page_icon="📊",
)

# =========================
# Global CSS
# =========================
st.markdown(
    """
<style>
.block-container {
    padding-top: 1.0rem;
    padding-bottom: 2rem;
    max-width: 1500px;
}

h1 {
    color: #1F4E78;
    font-weight: 800;
    letter-spacing: -0.5px;
    margin-bottom: 0.2rem;
}

h2, h3, h4 {
    color: #1F4E78;
}

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #f8fbff 0%, #eef5fb 100%);
    border-right: 1px solid #d9e2ec;
}

/* Buttons */
.stButton > button {
    border-radius: 12px;
    padding: 0.55rem 1rem;
    font-weight: 600;
    border: none;
    background: linear-gradient(90deg, #1F4E78 0%, #2E6FA3 100%);
    color: white;
}
.stButton > button:hover {
    opacity: 0.92;
    transform: translateY(-1px);
}

.stDownloadButton > button {
    border-radius: 12px;
    padding: 0.55rem 1rem;
    font-weight: 600;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 10px;
    border-bottom: 1px solid #e6edf3;
    padding-bottom: 2px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 14px 14px 0 0;
    padding: 0.7rem 1.1rem;
    background: #f7f9fc;
    font-weight: 600;
    color: #475569;
}
.stTabs [aria-selected="true"] {
    background: white;
    color: #1F4E78;
    border-bottom: 3px solid #1F4E78;
}

/* Card-like containers */
.control-card {
    background: white;
    border: 1px solid #e5edf5;
    border-radius: 18px;
    padding: 1rem 1rem 0.8rem 1rem;
    box-shadow: 0 4px 18px rgba(31, 78, 120, 0.05);
    margin-bottom: 1rem;
}

.section-card {
    background: white;
    border: 1px solid #e5edf5;
    border-radius: 18px;
    padding: 1rem;
    box-shadow: 0 4px 18px rgba(31, 78, 120, 0.05);
    margin-bottom: 1rem;
}

.section-title {
    font-size: 1.05rem;
    font-weight: 800;
    color: #1F4E78;
    margin-bottom: 0.15rem;
}

.section-subtitle {
    font-size: 0.87rem;
    color: #64748b;
    margin-bottom: 0.75rem;
}

.metric-card {
    background: white;
    border: 1px solid #e5edf5;
    padding: 1rem 1.1rem;
    border-radius: 16px;
    box-shadow: 0 2px 10px rgba(31, 78, 120, 0.06);
}
.metric-label {
    font-size: 0.85rem;
    color: #64748b;
}
.metric-value {
    font-size: 1.6rem;
    font-weight: 800;
    color: #1F4E78;
}
.metric-sub {
    font-size: 0.8rem;
    color: #94a3b8;
}

/* Data editor / dataframe */
div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    border-radius: 16px;
    overflow: hidden;
    border: 1px solid #e5edf5;
    box-shadow: 0 3px 14px rgba(31, 78, 120, 0.05);
    background: white;
}

/* Reduce grid feeling */
div[data-testid="stDataFrame"] table,
div[data-testid="stDataEditor"] table {
    border-collapse: separate !important;
    border-spacing: 0 !important;
}

div[data-testid="stDataFrame"] thead th,
div[data-testid="stDataEditor"] thead th {
    background: linear-gradient(180deg, #f8fbff 0%, #eef5fb 100%) !important;
    color: #1f2937 !important;
    font-weight: 700 !important;
    border-bottom: 1px solid #dce7f2 !important;
}

div[data-testid="stDataFrame"] tbody td,
div[data-testid="stDataEditor"] tbody td {
    border-bottom: 1px solid #f0f4f8 !important;
}

div[data-testid="stDataFrame"] tbody tr:nth-child(even),
div[data-testid="stDataEditor"] tbody tr:nth-child(even) {
    background-color: #fcfdff !important;
}

div[data-testid="stDataFrame"] tbody tr:hover,
div[data-testid="stDataEditor"] tbody tr:hover {
    background-color: #f4f9ff !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# =========================
# Header
# =========================
st.title("Output Simple Planner")
st.markdown(
    """
<div style="
    background: linear-gradient(90deg, #1F4E78 0%, #2E6FA3 100%);
    padding: 18px 20px;
    border-radius: 16px;
    color: white;
    margin-bottom: 18px;
    box-shadow: 0 4px 18px rgba(31,78,120,0.15);
">
    <div style="font-size: 1.1rem; font-weight: 700;">Interactive Planning Dashboard</div>
    <div style="opacity: 0.92; margin-top: 4px;">
        More control-panel style, less spreadsheet style.
    </div>
</div>
""",
    unsafe_allow_html=True,
)

# =========================
# Helpers
# =========================
def sheet_name(excel_file, target):
    lookup = {s.strip().lower(): s for s in excel_file.sheet_names}
    return lookup.get(target.strip().lower())


@st.cache_data(show_spinner=False)
def load_tables(file_bytes):
    if file_bytes:
        source = BytesIO(file_bytes)
    elif SAMPLE_INPUT.exists():
        source = SAMPLE_INPUT
    elif LOCAL_INPUT.exists():
        source = LOCAL_INPUT
    else:
        return None

    xl = pd.ExcelFile(source)

    def read(target, required=True):
        name = sheet_name(xl, target)
        if name is None:
            if required:
                raise ValueError(f"Missing sheet: {target}")
            return pd.DataFrame()
        return pd.read_excel(xl, sheet_name=name)

    return {
        "flow": read(planner.FLOW_SHEET),
        "product": read(planner.PRODUCT_SHEET),
        "demand": read(planner.DEMAND_SHEET),
        "inventory": read(planner.INVENTORY_SHEET),
        "target": read(planner.TARGET_SHEET, required=False),
    }


def target_value(target_df, key, default):
    if target_df.empty or not {"Columns", "Value"}.issubset(target_df.columns):
        return default
    matches = target_df[target_df["Columns"].astype(str).str.strip().str.lower() == key.lower()]
    if matches.empty:
        return default
    value = pd.to_numeric(matches.iloc[0]["Value"], errors="coerce")
    return default if pd.isna(value) else value


def build_input_workbook(flow_df, product_df, demand_df, inventory_df, target_reach, tester_number):
    target_df = pd.DataFrame({
        "Columns": ["Target Reach Level", "Tester Number"],
        "Value": [target_reach, tester_number],
    })
    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp_path = Path(temp.name)
    temp.close()
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        flow_df.to_excel(writer, sheet_name=planner.FLOW_SHEET, index=False)
        product_df.to_excel(writer, sheet_name=planner.PRODUCT_SHEET, index=False)
        demand_df.to_excel(writer, sheet_name=planner.DEMAND_SHEET, index=False)
        inventory_df.to_excel(writer, sheet_name=planner.INVENTORY_SHEET, index=False)
        target_df.to_excel(writer, sheet_name=planner.TARGET_SHEET, index=False)
    return temp_path


def build_month_product_table(monthly_summary_df, value_col, aggfunc="sum"):
    output_df = monthly_summary_df.pivot_table(
        index="Month",
        columns="Product_Key",
        values=value_col,
        aggfunc=aggfunc,
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    ).reset_index()
    output_df = output_df.rename(columns={"Month": "Row Labels"})
    if aggfunc == "mean":
        value_columns = [col for col in output_df.columns if col != "Row Labels"]
        output_df[value_columns] = output_df[value_columns].round(2)
    return output_df


def build_wafer_start_table(monthly_summary_df):
    table_df = monthly_summary_df.pivot_table(
        index=["Basic_Type", "Product_Key"],
        columns="Month",
        values="WaferStart",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    ).reset_index()
    return table_df.rename(columns={"Basic_Type": "Basic Type", "Product_Key": "Row Labels"})


def build_stage_output_table(monthly_summary_df):
    metric_options = [
        ("Bump_Wafer", "Bump"),
        ("Sort_Wafer", "Sort"),
        ("Testing_Wafer", "Sort / Testing"),
        ("DPS_Chip", "DPS"),
        ("DPS_Output", "DPS"),
        ("Output", "DPS"),
    ]
    metric_map = {col: label for col, label in metric_options if col in monthly_summary_df.columns}
    long_df = monthly_summary_df.melt(
        id_vars=["Basic_Type", "Product_Key", "Month"],
        value_vars=list(metric_map.keys()),
        var_name="Metric",
        value_name="Value",
    )
    long_df["Metric"] = long_df["Metric"].map(metric_map)
    table_df = long_df.pivot_table(
        index=["Basic_Type", "Product_Key", "Metric"],
        columns="Month",
        values="Value",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    ).reset_index()
    return table_df.rename(columns={"Basic_Type": "Basic Type", "Product_Key": "Row Labels"})


def build_graph_outputs(monthly_summary_df):
    stock_df = build_month_product_table(monthly_summary_df, "End_Stock", "sum")
    demand_df = build_month_product_table(monthly_summary_df, "Demand", "sum")
    month_rows = demand_df["Row Labels"].astype(str) != "Grand Total"
    stock_df["Next Month Demand"] = 0
    stock_df.loc[month_rows, "Next Month Demand"] = demand_df.loc[month_rows, "Grand Total"].shift(-1).fillna(0).values
    return [
        ("Tester Used", build_month_product_table(monthly_summary_df, "Max_TesterUsed", "sum"), "bar"),
        ("VRFN Demand", build_month_product_table(monthly_summary_df, "Demand", "sum"), "bar"),
        ("Reach Level", build_month_product_table(monthly_summary_df, "Avg_REACH", "mean"), "line"),
        ("Stock Development", stock_df, "stock"),
    ]


def style_excel_range(worksheet, header_row, last_row, last_col):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="EAF3F8")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin_side = Side(style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in worksheet[header_row]:
        if cell.column <= last_col:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, max_col=last_col):
        is_total_row = any(str(cell.value) == "Grand Total" for cell in row[:3])
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_total_row:
                cell.fill = total_fill
                cell.font = bold_font
            elif cell.column <= 3:
                cell.fill = metric_fill

    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=2)

    for column_index in range(1, last_col + 1):
        max_len = max(
            len(str(worksheet.cell(row=row_index, column=column_index).value or ""))
            for row_index in range(header_row, last_row + 1)
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 11), 18)


def add_table_block(worksheet, title, output_df, start_row):
    worksheet.cell(row=start_row, column=1, value=title)
    worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14, color="1F4E78")
    header_row = start_row + 1
    for row_index, row_values in enumerate(dataframe_to_rows(output_df, index=False, header=True), header_row):
        for column_index, value in enumerate(row_values, 1):
            worksheet.cell(row=row_index, column=column_index, value=value)
    style_excel_range(worksheet, header_row, header_row + len(output_df), len(output_df.columns))
    return header_row, header_row + len(output_df), len(output_df.columns)


def add_excel_chart(worksheet, title, chart_kind, header_row, last_row, last_col, anchor):
    headers = [worksheet.cell(row=header_row, column=col).value for col in range(1, last_col + 1)]
    overlay_col = headers.index("Next Month Demand") + 1 if "Next Month Demand" in headers else None
    chart_last_col = overlay_col - 1 if overlay_col else last_col
    if worksheet.cell(row=header_row, column=chart_last_col).value == "Grand Total":
        chart_last_col -= 1
    if chart_last_col < 2 or last_row <= header_row:
        return

    chart_map = {"bar": BarChart, "line": LineChart, "area": AreaChart, "stock": BarChart}
    chart = chart_map[chart_kind]()
    chart.title = title
    chart.height = 10
    chart.width = 20
    data = Reference(worksheet, min_col=2, max_col=chart_last_col, min_row=header_row, max_row=last_row)
    categories = Reference(worksheet, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    if chart_kind in ["bar", "stock"]:
        chart.type = "col"
        chart.style = 10
    elif chart_kind == "area":
        chart.grouping = "stacked"
    if chart_kind == "stock" and overlay_col:
        line_chart = LineChart()
        line_data = Reference(worksheet, min_col=overlay_col, max_col=overlay_col, min_row=header_row, max_row=last_row)
        line_chart.add_data(line_data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.y_axis.axId = 200
        line_chart.y_axis.title = "Next Month Demand"
        chart += line_chart
    worksheet.add_chart(chart, anchor)


def style_excel_sheet(worksheet):
    style_excel_range(worksheet, 1, worksheet.max_row, worksheet.max_column)


def make_output_bytes(graph_outputs, wafer_start_table, stage_output_table):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wafer_start_table.to_excel(writer, sheet_name="Wafer_Start", index=False)
        stage_output_table.to_excel(writer, sheet_name="Bump_Testing_DPS", index=False)
        style_excel_sheet(writer.book["Wafer_Start"])
        style_excel_sheet(writer.book["Bump_Testing_DPS"])

        graph_sheet = writer.book.create_sheet("Graph", 0)
        for start_row, (title, graph_df, chart_kind) in zip([1, 30, 59, 88], graph_outputs):
            header_row, last_row, last_col = add_table_block(graph_sheet, title, graph_df, start_row)
            add_excel_chart(graph_sheet, title, chart_kind, header_row, last_row, last_col, f"L{start_row}")

    output.seek(0)
    return output.getvalue()


def style_display_table(output_df):
    numeric_cols = output_df.select_dtypes(include="number").columns.tolist()

    def heatmap_column(column):
        values = pd.to_numeric(column, errors="coerce")
        min_value = values.min()
        max_value = values.max()
        if pd.isna(min_value) or pd.isna(max_value) or min_value == max_value:
            return ["background-color: #f7fbff" for _ in column]
        styles = []
        for value in values:
            if pd.isna(value):
                styles.append("")
                continue
            strength = (value - min_value) / (max_value - min_value)
            blue = int(245 - 55 * strength)
            green = int(250 - 80 * strength)
            red = int(255 - 120 * strength)
            styles.append(f"background-color: rgb({red}, {green}, {blue})")
        return styles

    def highlight_total(row):
        is_total = any(str(value) == "Grand Total" for value in row.values)
        return ["background-color: #d9eaf7; font-weight: 700" if is_total else "" for _ in row]

    styler = (
        output_df.style
        .format({col: "{:,.0f}" for col in numeric_cols})
        .set_table_styles([
            {"selector": "th", "props": [("background-color", "#f8fbff"), ("color", "#1f2937"), ("font-weight", "700")]},
            {"selector": "td", "props": [("border-color", "#edf2f7")]},
        ])
    )
    if numeric_cols:
        styler = styler.apply(heatmap_column, subset=numeric_cols, axis=0)
    return styler.apply(highlight_total, axis=1)


def render_table_card(title, subtitle, styled_df, height=430):
    st.markdown(f"<div class='section-title'>{title}</div>", unsafe_allow_html=True)
    if subtitle:
        st.markdown(f"<div class='section-subtitle'>{subtitle}</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-card'>", unsafe_allow_html=True)
    st.dataframe(styled_df, width="stretch", height=height)
    st.markdown("</div>", unsafe_allow_html=True)


def draw_graphs(graph_outputs):
    st.subheader("Graphs")
    for title, graph_df, chart_kind in graph_outputs:
        chart_df = graph_df[graph_df["Row Labels"].astype(str) != "Grand Total"].copy()

        if chart_kind == "stock":
            demand_line = chart_df[["Row Labels", "Next Month Demand"]].copy()
            product_cols = [col for col in chart_df.columns if col not in ["Row Labels", "Grand Total", "Next Month Demand"]]
            chart_df = chart_df.melt(id_vars="Row Labels", value_vars=product_cols, var_name="Product", value_name=title)
            chart = px.bar(chart_df, x="Row Labels", y=title, color="Product", title=title, template="plotly_white")
            chart.update_layout(barmode="stack")
            chart.add_scatter(
                x=demand_line["Row Labels"],
                y=demand_line["Next Month Demand"],
                mode="lines+markers",
                name="Next Month Demand",
                line={"color": "#d62728", "width": 3},
                marker={"size": 8},
            )
            chart.update_layout(xaxis_title="Month", yaxis_title="Stock / Next Month Demand", margin=dict(l=20, r=20, t=50, b=20))
            st.plotly_chart(chart, use_container_width=True)
            continue

        chart_df = chart_df.melt(id_vars="Row Labels", var_name="Product", value_name=title)
        chart_df = chart_df[chart_df["Product"] != "Grand Total"]

        if chart_kind == "line":
            chart = px.line(chart_df, x="Row Labels", y=title, color="Product", markers=True, title=title, template="plotly_white")
        else:
            chart = px.bar(chart_df, x="Row Labels", y=title, color="Product", title=title, template="plotly_white")
            chart.update_layout(barmode="stack")

        chart.update_layout(xaxis_title="Month", yaxis_title=title, margin=dict(l=20, r=20, t=50, b=20))
        st.plotly_chart(chart, use_container_width=True)


def text_value(v):
    return "" if pd.isna(v) else v


# =========================
# Load data
# =========================
uploaded = st.sidebar.file_uploader("Upload input Excel", type=["xlsx"])
file_bytes = uploaded.getvalue() if uploaded else None

try:
    tables = load_tables(file_bytes)
except Exception as exc:
    st.error(str(exc))
    st.stop()

if tables is None:
    st.info("Please upload an input Excel file from the sidebar, or add input_sample.xlsx to the repo.")
    st.stop()

target_reach_default = float(target_value(tables["target"], "Target Reach Level", planner.DEFAULT_TESTER_CONFIG["target_REACH"]))
tester_number_default = int(target_value(tables["target"], "Tester Number", planner.DEFAULT_TESTER_CONFIG["available"]))

# =========================
# Top control panel
# =========================
st.markdown("<div class='control-card'>", unsafe_allow_html=True)
st.markdown("<div class='section-title'>Global Planning Controls</div>", unsafe_allow_html=True)
st.markdown("<div class='section-subtitle'>Adjust key parameters first. Then edit inputs below.</div>", unsafe_allow_html=True)

c1, c2 = st.columns(2)
with c1:
    target_reach = st.number_input("Target Reach Level", min_value=0.0, value=target_reach_default, step=0.5)
with c2:
    tester_number = st.number_input("Tester Number", min_value=1, value=tester_number_default, step=1)

st.markdown("</div>", unsafe_allow_html=True)

# =========================
# Input mode
# =========================
edit_mode = st.radio("Input editor style", ["Card view", "Table view"], horizontal=True)

# =========================
# Edit flow / product / demand / inventory
# =========================
def edit_row_card(df, i, row, field_specs, prefix):
    cols = st.columns([1] * len(field_specs))
    for j, spec in enumerate(field_specs):
        col_name = spec["col"]
        label = spec.get("label", col_name)
        kind = spec.get("kind", "text")
        default = text_value(row[col_name])

        with cols[j]:
            if kind == "number":
                if pd.isna(default) or default == "":
                    default = 0.0
                df.at[i, col_name] = st.number_input(
                    label,
                    value=float(default),
                    key=f"{prefix}_{i}_{col_name}",
                )
            else:
                df.at[i, col_name] = st.text_input(
                    label,
                    value=str(default),
                    key=f"{prefix}_{i}_{col_name}",
                )


tabs = st.tabs(["Flow", "Product", "Demand", "Inventory"])

with tabs[0]:
    st.markdown("<div class='section-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Flow</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-subtitle'>Edit stages and processing parameters.</div>", unsafe_allow_html=True)

    if edit_mode == "Card view":
        flow_df = tables["flow"].copy()
        flow_fields = [
            {"col": flow_df.columns[0], "label": flow_df.columns[0], "kind": "text"},
            {"col": flow_df.columns[1], "label": flow_df.columns[1], "kind": "number"},
            {"col": flow_df.columns[2], "label": flow_df.columns[2], "kind": "number"},
            {"col": flow_df.columns[3], "label": flow_df.columns[3], "kind": "number"},
        ]
        for i, row in flow_df.iterrows():
            with st.container(border=True):
                st.markdown(f"**Step {i+1}**")
                edit_row_card(flow_df, i, row, flow_fields, "flow")
        st.caption("Tip: card view reduces spreadsheet feeling and improves readability.")
    else:
        flow_df = st.data_editor(tables["flow"], width="stretch", num_rows="dynamic", key="simple_flow", hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

with tabs[1]:
    st.markdown("<div class='section-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Product</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-subtitle'>Edit product master data.</div>", unsafe_allow_html=True)

    if edit_mode == "Card view":
        product_df = tables["product"].copy()
        # fallback to generic editable fields
        field_specs = []
        for col in product_df.columns[:4]:
            if pd.api.types.is_numeric_dtype(product_df[col]):
                field_specs.append({"col": col, "label": col, "kind": "number"})
            else:
                field_specs.append({"col": col, "label": col, "kind": "text"})

        for i, row in product_df.iterrows():
            with st.container(border=True):
                st.markdown(f"**Product {i+1}**")
                edit_row_card(product_df, i, row, field_specs, "product")
    else:
        product_df = st.data_editor(tables["product"], width="stretch", num_rows="dynamic", key="simple_product", hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

with tabs[2]:
    st.markdown("<div class='section-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Demand</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-subtitle'>Adjust demand assumptions.</div>", unsafe_allow_html=True)

    if edit_mode == "Card view":
        demand_df = st.data_editor(tables["demand"], width="stretch", num_rows="dynamic", key="simple_demand", hide_index=True)
        st.caption("For demand data with many columns, table view may still be more practical.")
    else:
        demand_df = st.data_editor(tables["demand"], width="stretch", num_rows="dynamic", key="simple_demand", hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

with tabs[3]:
    st.markdown("<div class='section-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Inventory</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-subtitle'>Update inventory assumptions and starting stock.</div>", unsafe_allow_html=True)

    if edit_mode == "Card view":
        inventory_df = st.data_editor(tables["inventory"], width="stretch", num_rows="dynamic", key="simple_inventory", hide_index=True)
        st.caption("Inventory often has many columns, so the table component is kept for usability.")
    else:
        inventory_df = st.data_editor(tables["inventory"], width="stretch", num_rows="dynamic", key="simple_inventory", hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

# =========================
# Run planner
# =========================
if st.button("Run simple plan", type="primary"):
    with st.spinner("Running simple reach-based plan..."):
        try:
            input_path = build_input_workbook(
                flow_df, product_df, demand_df, inventory_df,
                target_reach, tester_number
            )
            stages, products, weekly_demand_map, month_label_map, tester_config, inventory_map = planner.load_all_inputs(input_path)
            results = planner.run_all_products(stages, products, weekly_demand_map, month_label_map, tester_config, inventory_map)
        except Exception as exc:
            st.error(str(exc))
            st.stop()

    all_plan_df, summary_df, monthly_summary_df, skipped_df = results
    graph_outputs = build_graph_outputs(monthly_summary_df)
    wafer_start_table = build_wafer_start_table(monthly_summary_df)
    stage_output_table = build_stage_output_table(monthly_summary_df)
    output_bytes = make_output_bytes(graph_outputs, wafer_start_table, stage_output_table)

    st.success("Plan generated successfully")

    total_products = len(product_df)
    total_demand = pd.to_numeric(demand_df.select_dtypes(include="number").sum().sum(), errors="coerce")
    total_output = pd.to_numeric(monthly_summary_df.get("Output", pd.Series(dtype=float)).sum(), errors="coerce")
    skipped_count = 0 if skipped_df.empty else len(skipped_df)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Products</div>
            <div class="metric-value">{total_products}</div>
            <div class="metric-sub">Editable items</div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total Demand</div>
            <div class="metric-value">{total_demand:,.0f}</div>
            <div class="metric-sub">From demand sheet</div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Total Output</div>
            <div class="metric-value">{total_output:,.0f}</div>
            <div class="metric-sub">Planner result</div>
        </div>
        """, unsafe_allow_html=True)
    with c4:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Skipped Products</div>
            <div class="metric-value">{skipped_count}</div>
            <div class="metric-sub">Not planned</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

    draw_graphs(graph_outputs)

    st.subheader("Output Tables")
    render_table_card("Wafer Start", "Monthly wafer start summary by product.", style_display_table(wafer_start_table), height=430)
    render_table_card("Bump Testing DPS", "Stage output summary for bump / testing / DPS.", style_display_table(stage_output_table), height=430)

    with st.expander("Summary", expanded=False):
        st.dataframe(summary_df, width="stretch", height=400)

    if not skipped_df.empty:
        with st.expander("Skipped Products", expanded=False):
            st.dataframe(skipped_df, width="stretch", height=300)

    st.download_button(
        "Download Simple Output Excel",
        data=output_bytes,
        file_name="output_simple_streamlit_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
