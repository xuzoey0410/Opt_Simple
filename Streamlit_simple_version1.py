from io import BytesIO
import importlib
from pathlib import Path
import sys
import tempfile

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

APP_DIR = Path(__file__).parent
SAMPLE_INPUT = APP_DIR / "input_sample.xlsx"
LOCAL_INPUT = APP_DIR / "All_Input1.xlsx"
LOCAL_OLD_INPUT = APP_DIR / "All_Output1.xlsx"
PROJECT_DIR = Path.home() / "Desktop" / "Demand Project"
PROJECT_SAMPLE_INPUT = PROJECT_DIR / "input_sample.xlsx"
PROJECT_LOCAL_INPUT = PROJECT_DIR / "All_Input1.xlsx"
PROJECT_OLD_INPUT = PROJECT_DIR / "All_Output1.xlsx"

for candidate_dir in [APP_DIR, PROJECT_DIR]:
    candidate_path = str(candidate_dir)
    sys.path = [path for path in sys.path if path != candidate_path]

for candidate_dir in [APP_DIR, PROJECT_DIR]:
    if candidate_dir.exists():
        sys.path.insert(0, str(candidate_dir))

if "Output_Simple" in sys.modules:
    del sys.modules["Output_Simple"]
planner = importlib.import_module("Output_Simple")

st.set_page_config(page_title="Supply Chain Planner", layout="wide")
st.title("Supply Chain Planner")


def sheet_name(excel_file, target):
    lookup = {s.strip().lower(): s for s in excel_file.sheet_names}
    return lookup.get(target.strip().lower())


@st.cache_data(show_spinner=False)
def load_tables(file_bytes):
    if file_bytes:
        source = BytesIO(file_bytes)
    elif SAMPLE_INPUT.exists():
        source = SAMPLE_INPUT
    elif LOCAL_INPUT.exists():
        source = LOCAL_INPUT
    elif LOCAL_OLD_INPUT.exists():
        source = LOCAL_OLD_INPUT
    elif PROJECT_SAMPLE_INPUT.exists():
        source = PROJECT_SAMPLE_INPUT
    elif PROJECT_LOCAL_INPUT.exists():
        source = PROJECT_LOCAL_INPUT
    elif PROJECT_OLD_INPUT.exists():
        source = PROJECT_OLD_INPUT
    else:
        return None

    xl = pd.ExcelFile(source)

    def read(target, required=True):
        name = sheet_name(xl, target)
        if name is None:
            if required:
                raise ValueError(f"Missing sheet: {target}")
            return pd.DataFrame()
        return pd.read_excel(xl, sheet_name=name)

    return {
        "flow": read(planner.FLOW_SHEET),
        "product": read(planner.PRODUCT_SHEET),
        "demand": read(planner.DEMAND_SHEET),
        "inventory": read(planner.INVENTORY_SHEET),
        "target": read(planner.TARGET_SHEET, required=False),
    }


def target_value(target_df, key, default):
    if target_df.empty or not {"Columns", "Value"}.issubset(target_df.columns):
        return default
    matches = target_df[target_df["Columns"].astype(str).str.strip().str.lower() == key.lower()]
    if matches.empty:
        return default
    value = pd.to_numeric(matches.iloc[0]["Value"], errors="coerce")
    return default if pd.isna(value) else value


def ensure_priority_column(product_df):
    if any(str(col).strip().lower() == "priority" for col in product_df.columns):
        return product_df

    updated_df = product_df.copy()
    updated_df.insert(len(updated_df.columns), "Priority", 0)
    return updated_df


def demand_month_columns(demand_df):
    return [col for col in demand_df.columns if planner.is_month_col(col)]


def normalize_month_column(value):
    raw_value = str(value).strip()
    if not raw_value:
        raise ValueError("Please enter a demand month, for example 2028-10.")

    parsed = pd.to_datetime(raw_value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError("Month format is not recognized. Please use a format like 2028-10 or Oct-2028.")
    return parsed.strftime("%Y-%m")


def add_demand_month(demand_df, month_value):
    month_col = normalize_month_column(month_value)
    existing_cols = {str(col).strip(): col for col in demand_df.columns}
    if month_col in existing_cols:
        raise ValueError(f"Demand month already exists: {month_col}")

    month_cols = demand_month_columns(demand_df)
    insert_at = demand_df.columns.get_loc(month_cols[-1]) + 1 if month_cols else len(demand_df.columns)
    updated_df = demand_df.copy()
    updated_df.insert(insert_at, month_col, 0)
    return updated_df, month_col


def demand_date_range_defaults(demand_df):
    month_cols = demand_month_columns(demand_df)
    if not month_cols:
        today = pd.Timestamp.today().replace(day=1)
        return today.date(), (today + pd.DateOffset(months=5)).date()

    parsed_months = pd.to_datetime([normalize_month_column(col) for col in month_cols], errors="coerce")
    parsed_months = [month for month in parsed_months if not pd.isna(month)]
    if not parsed_months:
        today = pd.Timestamp.today().replace(day=1)
        return today.date(), (today + pd.DateOffset(months=5)).date()

    return min(parsed_months).date(), max(parsed_months).date()


def apply_demand_date_range(demand_df, start_value, end_value):
    start_month = pd.to_datetime(start_value, errors="coerce")
    end_month = pd.to_datetime(end_value, errors="coerce")
    if pd.isna(start_month) or pd.isna(end_month):
        raise ValueError("Please select both start and end month.")

    start_month = start_month.replace(day=1)
    end_month = end_month.replace(day=1)
    if start_month > end_month:
        raise ValueError("Start month must be before or equal to end month.")

    updated_df = demand_df.copy()
    old_month_cols = demand_month_columns(updated_df)
    old_month_lookup = {normalize_month_column(col): col for col in old_month_cols}
    id_cols = [col for col in updated_df.columns if col not in old_month_cols]
    new_month_cols = [month.strftime("%Y-%m") for month in pd.date_range(start_month, end_month, freq="MS")]

    output_df = updated_df[id_cols].copy()
    for month_col in new_month_cols:
        if month_col in old_month_lookup:
            output_df[month_col] = updated_df[old_month_lookup[month_col]]
        else:
            output_df[month_col] = 0

    return output_df


def sync_header_labels(table_key, columns, source_key):
    labels_key = f"{table_key}_header_labels"
    source_state_key = f"{table_key}_header_source"
    columns = [str(col) for col in columns]

    if st.session_state.get(source_state_key) != source_key or labels_key not in st.session_state:
        st.session_state[source_state_key] = source_key
        st.session_state[labels_key] = {col: col for col in columns}
    else:
        labels = st.session_state[labels_key]
        st.session_state[labels_key] = {col: labels.get(col, col) for col in columns}

    return st.session_state[labels_key]


def render_header_editor(table_key, columns, source_key):
    columns = [str(col) for col in columns]
    labels = sync_header_labels(table_key, columns, source_key)
    editor_df = pd.DataFrame({
        "Column": columns,
        "Display Header": [labels.get(col, col) for col in columns],
    })

    with st.expander("Edit table headers"):
        edited_df = st.data_editor(
            editor_df,
            width="stretch",
            hide_index=True,
            key=f"{table_key}_header_editor",
            column_config={
                "Column": st.column_config.TextColumn("Column", disabled=True),
                "Display Header": st.column_config.TextColumn("Display Header"),
            },
        )

        if st.button("Apply headers", key=f"{table_key}_apply_headers"):
            new_labels = {}
            for _, row in edited_df.iterrows():
                column = str(row.get("Column", "")).strip()
                label = str(row.get("Display Header", "")).strip()
                if column:
                    new_labels[column] = label or column
            st.session_state[f"{table_key}_header_labels"] = {col: new_labels.get(col, col) for col in columns}
            st.success("Headers updated")

    return st.session_state[f"{table_key}_header_labels"]


def generic_column_config(columns, labels):
    return {col: st.column_config.Column(label=labels.get(str(col), str(col))) for col in columns}


def flow_to_editor_table(flow_df):
    df = planner.clean_columns(flow_df).dropna(how="all")
    if df.empty:
        return pd.DataFrame(columns=["Stage", "Cycle Time Week", "Transit Week", "Yield"])

    stage_col = planner.find_col(df, ["Stages", "Stage", "Process", "Flow"])
    week_col = planner.find_col(df, ["Time/Week", "Time Week", "Cycle_Time_Week", "Cycle Time Week"])
    yield_col = planner.find_col(df, ["Yield", "Yield Rate"], required=False)

    rows = []
    for _, row in df.iterrows():
        stage_value = row.get(stage_col)
        if pd.isna(stage_value):
            continue

        stage_name = str(stage_value).strip()
        stage_lower = stage_name.lower()
        week_value = pd.to_numeric(row.get(week_col), errors="coerce")
        week_value = 0 if pd.isna(week_value) else float(week_value)

        if "transit" in stage_lower or "transist" in stage_lower:
            if rows:
                rows[-1]["Transit Week"] += week_value
            continue

        yield_value = pd.to_numeric(row.get(yield_col), errors="coerce") if yield_col else 1
        rows.append({
            "Stage": stage_name,
            "Cycle Time Week": week_value,
            "Transit Week": 0.0,
            "Yield": 1.0 if pd.isna(yield_value) else float(yield_value),
        })

    return pd.DataFrame(rows, columns=["Stage", "Cycle Time Week", "Transit Week", "Yield"])


def flow_editor_to_workbook_table(flow_df):
    if not {"Stage", "Cycle Time Week", "Transit Week"}.issubset(flow_df.columns):
        return flow_df

    rows = []
    for _, row in flow_df.dropna(how="all").iterrows():
        stage_name = str(row.get("Stage", "")).strip()
        if not stage_name:
            continue

        cycle_week = pd.to_numeric(row.get("Cycle Time Week"), errors="coerce")
        transit_week = pd.to_numeric(row.get("Transit Week"), errors="coerce")
        yield_value = pd.to_numeric(row.get("Yield"), errors="coerce")

        rows.append({
            "Stages": stage_name,
            "Time/Week": 0 if pd.isna(cycle_week) else cycle_week,
            "Yield": 1 if pd.isna(yield_value) else yield_value,
        })

        if not pd.isna(transit_week) and transit_week > 0:
            rows.append({
                "Stages": "Transist Time",
                "Time/Week": transit_week,
                "Yield": None,
            })

    return pd.DataFrame(rows, columns=["Stages", "Time/Week", "Yield"])


def build_input_workbook(flow_df, product_df, demand_df, inventory_df, min_reach, target_reach, max_reach, tester_number, tester_smoothing_weeks, wafer_start_before):
    target_df = pd.DataFrame({
        "Columns": ["Min Reach Level", "Target Reach Level", "Max Reach Level", "Tester Number", "Tester Smoothing Weeks", "Wafer Start Time"],
        "Value": [min_reach, target_reach, max_reach, tester_number, tester_smoothing_weeks, wafer_start_before],
    })

    flow_output_df = flow_editor_to_workbook_table(flow_df)

    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp_path = Path(temp.name)
    temp.close()

    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        flow_output_df.to_excel(writer, sheet_name=planner.FLOW_SHEET, index=False)
        product_df.to_excel(writer, sheet_name=planner.PRODUCT_SHEET, index=False)
        demand_df.to_excel(writer, sheet_name=planner.DEMAND_SHEET, index=False)
        inventory_df.to_excel(writer, sheet_name=planner.INVENTORY_SHEET, index=False)
        target_df.to_excel(writer, sheet_name=planner.TARGET_SHEET, index=False)

    return temp_path


def make_input_bytes(flow_df, product_df, demand_df, inventory_df, min_reach, target_reach, max_reach, tester_number, tester_smoothing_weeks, wafer_start_before):
    target_df = pd.DataFrame({
        "Columns": ["Min Reach Level", "Target Reach Level", "Max Reach Level", "Tester Number", "Tester Smoothing Weeks", "Wafer Start Time"],
        "Value": [min_reach, target_reach, max_reach, tester_number, tester_smoothing_weeks, wafer_start_before],
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        flow_editor_to_workbook_table(flow_df).to_excel(writer, sheet_name=planner.FLOW_SHEET, index=False)
        product_df.to_excel(writer, sheet_name=planner.PRODUCT_SHEET, index=False)
        demand_df.to_excel(writer, sheet_name=planner.DEMAND_SHEET, index=False)
        inventory_df.to_excel(writer, sheet_name=planner.INVENTORY_SHEET, index=False)
        target_df.to_excel(writer, sheet_name=planner.TARGET_SHEET, index=False)
    output.seek(0)
    return output.getvalue()


def parse_wafer_start_before(value):
    if not str(value).strip():
        return None

    wafer_start_before = int(float(str(value).strip()))
    if wafer_start_before < 0:
        raise ValueError("Wafer Start Time must be blank or a non-negative number of weeks.")
    return wafer_start_before


def build_month_product_table(monthly_summary_df, value_col, aggfunc="sum"):
    month_order = list(dict.fromkeys(monthly_summary_df["Month"].tolist()))
    output_df = monthly_summary_df.pivot_table(
        index="Month",
        columns="Product_Key",
        values=value_col,
        aggfunc=aggfunc,
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    ).reset_index()

    output_df = output_df.rename(columns={"Month": "Row Labels"})
    row_order = {month: index for index, month in enumerate(month_order)}
    output_df["_Month_Order"] = output_df["Row Labels"].map(lambda value: row_order.get(value, len(row_order)))
    output_df = output_df.sort_values("_Month_Order", kind="stable").drop(columns="_Month_Order")
    if aggfunc == "mean":
        value_columns = [col for col in output_df.columns if col != "Row Labels"]
        output_df[value_columns] = output_df[value_columns].round(2)
    return output_df


def reorder_month_columns(output_df, id_columns, month_order):
    ordered_month_columns = [month for month in month_order if month in output_df.columns]
    other_columns = [
        col for col in output_df.columns
        if col not in id_columns and col not in ordered_month_columns and col != "Grand Total"
    ]
    kept_columns = id_columns + ordered_month_columns + other_columns + [col for col in ["Grand Total"] if col in output_df.columns]
    return output_df[kept_columns].copy()


def trim_leading_zero_month_columns(output_df, id_columns):
    month_columns = [col for col in output_df.columns if col not in id_columns and col != "Grand Total"]
    first_nonzero_index = None

    for index, col in enumerate(month_columns):
        values = pd.to_numeric(output_df[col], errors="coerce").fillna(0)
        if values.abs().sum() > 0:
            first_nonzero_index = index
            break

    if first_nonzero_index is None:
        return output_df[id_columns + [col for col in ["Grand Total"] if col in output_df.columns]].copy()

    kept_month_columns = month_columns[first_nonzero_index:]
    kept_columns = id_columns + kept_month_columns + [col for col in ["Grand Total"] if col in output_df.columns]
    return output_df[kept_columns].copy()


def trim_leading_zero_rows(output_df, value_columns):
    if output_df.empty:
        return output_df

    totals = output_df[value_columns].apply(pd.to_numeric, errors="coerce").fillna(0).abs().sum(axis=1)
    nonzero_positions = totals[totals > 0].index
    if len(nonzero_positions) == 0:
        return output_df.iloc[0:0].copy()
    return output_df.loc[nonzero_positions[0]:].copy()


def build_wafer_start_table(monthly_summary_df):
    month_order = list(dict.fromkeys(monthly_summary_df["Month"].tolist()))
    table_df = monthly_summary_df.pivot_table(
        index=["Basic_Type", "Product_Key"],
        columns="Month",
        values="WaferStart",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="Grand Total",
    ).reset_index()

    table_df = table_df.rename(columns={"Basic_Type": "Basic Type", "Product_Key": "Row Labels"})
    table_df = reorder_month_columns(table_df, ["Basic Type", "Row Labels"], month_order)
    return trim_leading_zero_month_columns(table_df, ["Basic Type", "Row Labels"])


def build_stage_output_table(monthly_summary_df):
    month_order = list(dict.fromkeys(monthly_summary_df["Month"].tolist()))
    metric_options = [
        ("Bump_Wafer", "Bump"),
        ("Sort_Wafer", "Testing"),
        ("Testing_Wafer", "Testing"),
        ("DPS_Chip", "DPS"),
        ("DPS_Output", "DPS"),
        ("Output", "DPS"),
    ]
    metric_map = {col: label for col, label in metric_options if col in monthly_summary_df.columns}
    metric_order = {"Bump": 1, "Testing": 2, "DPS": 3}

    long_df = monthly_summary_df.melt(
        id_vars=["Basic_Type", "Product_Key", "Month"],
        value_vars=list(metric_map.keys()),
        var_name="Metric",
        value_name="Value",
    )
    long_df["Metric"] = long_df["Metric"].map(metric_map)
    long_df["Metric_Order"] = long_df["Metric"].map(metric_order)

    table_df = long_df.pivot_table(
        index=["Metric_Order", "Metric", "Basic_Type", "Product_Key"],
        columns="Month",
        values="Value",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    table_df = table_df.sort_values(["Metric_Order", "Basic_Type", "Product_Key"], kind="stable")
    month_columns = [col for col in table_df.columns if col not in ["Metric_Order", "Metric", "Basic_Type", "Product_Key"]]
    total_row = {
        "Metric_Order": 99,
        "Metric": "Grand Total",
        "Basic_Type": "Grand Total",
        "Product_Key": "Grand Total",
    }
    dps_rows = table_df[table_df["Metric"].eq("DPS")]
    total_row.update({col: dps_rows[col].sum() for col in month_columns})
    table_df = pd.concat([table_df, pd.DataFrame([total_row])], ignore_index=True)
    table_df = table_df.drop(columns="Metric_Order")
    table_df = table_df.rename(columns={"Basic_Type": "Basic Type", "Product_Key": "Row Labels"})
    table_df = reorder_month_columns(table_df, ["Metric", "Basic Type", "Row Labels"], month_order)
    return trim_leading_zero_month_columns(table_df, ["Metric", "Basic Type", "Row Labels"])


def calc_target_stock(month_labels, demand_values, target_reach):
    target_values = []
    if not month_labels or not demand_values:
        return target_values

    for index in range(len(demand_values)):
        remaining_weeks = float(target_reach)
        target_stock = 0.0
        future_index = index + 1

        while remaining_weeks > 0:
            demand_index = min(future_index, len(demand_values) - 1)
            month_weeks = planner.month_to_weeks(month_labels[demand_index])
            if month_weeks <= 0:
                future_index += 1
                continue

            covered_weeks = min(remaining_weeks, month_weeks)
            weekly_demand = float(demand_values[demand_index]) / month_weeks
            target_stock += weekly_demand * covered_weeks
            remaining_weeks -= covered_weeks
            future_index += 1

        target_values.append(target_stock)

    return target_values


def build_tester_usage_table(all_plan_df, tester_capacity):
    tester_df = all_plan_df.copy()
    tester_df["TesterUsed"] = pd.to_numeric(tester_df["TesterUsed"], errors="coerce").fillna(0)
    weekly_df = (
        tester_df
        .groupby("Week", sort=True)["TesterUsed"]
        .sum()
        .reset_index(name="Total Tester Used")
    )
    weekly_df["Tester Capacity"] = int(tester_capacity)
    weekly_df["Remaining Tester"] = weekly_df["Tester Capacity"] - weekly_df["Total Tester Used"]
    weekly_df["Tester OK"] = weekly_df["Total Tester Used"] <= weekly_df["Tester Capacity"]
    return weekly_df


def build_tester_graph_table(all_plan_df, tester_capacity):
    tester_df = all_plan_df.copy()
    tester_df["TesterUsed"] = pd.to_numeric(tester_df["TesterUsed"], errors="coerce").fillna(0)
    weekly_product_df = (
        tester_df
        .pivot_table(index="Week", columns="Product_Key", values="TesterUsed", aggfunc="sum", fill_value=0)
        .reset_index()
        .rename(columns={"Week": "Row Labels"})
    )
    product_columns = [col for col in weekly_product_df.columns if col != "Row Labels"]
    weekly_product_df["Grand Total"] = weekly_product_df[product_columns].sum(axis=1)
    weekly_product_df["Tester Capacity"] = int(tester_capacity)
    return trim_leading_zero_rows(weekly_product_df, product_columns)


def build_unconstrained_tester_requirement_table(unconstrained_plan_df, tester_capacity):
    weekly_df = build_tester_usage_table(unconstrained_plan_df, tester_capacity)
    weekly_df = weekly_df.rename(columns={
        "Total Tester Used": "Required Tester No Capacity",
        "Remaining Tester": "Shortage vs Capacity",
        "Tester OK": "Within Current Capacity",
    })
    weekly_df["Shortage vs Capacity"] = (
        weekly_df["Required Tester No Capacity"] - weekly_df["Tester Capacity"]
    ).clip(lower=0)
    weekly_df["Within Current Capacity"] = weekly_df["Required Tester No Capacity"] <= weekly_df["Tester Capacity"]
    return weekly_df


def build_graph_outputs(monthly_summary_df, all_plan_df, target_reach, tester_capacity):
    graph_monthly_summary_df = monthly_summary_df[pd.to_numeric(monthly_summary_df["Demand"], errors="coerce").fillna(0) > 0].copy()
    stock_df = build_month_product_table(graph_monthly_summary_df, "End_Stock", "sum")
    demand_df = build_month_product_table(graph_monthly_summary_df, "Demand", "sum")
    month_rows = demand_df["Row Labels"].apply(planner.is_month_col)
    stock_df["Target Stock"] = 0.0
    month_labels = demand_df.loc[month_rows, "Row Labels"].tolist()
    demand_values = demand_df.loc[month_rows, "Grand Total"].astype(float).tolist()
    real_month_labels = stock_df["Row Labels"].apply(planner.is_month_col)
    stock_df.loc[real_month_labels, "Target Stock"] = calc_target_stock(month_labels, demand_values, target_reach)

    return [
        ("Tester Used", build_tester_graph_table(all_plan_df, tester_capacity), "tester"),
        ("VRFN Demand", build_month_product_table(graph_monthly_summary_df, "Demand", "sum"), "bar"),
        ("Reach Level", build_month_product_table(graph_monthly_summary_df, "Avg_REACH", "mean"), "line"),
        ("Stock Development", stock_df, "stock"),
    ]


def style_excel_range(worksheet, header_row, last_row, last_col, freeze=True):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="EAF3F8")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin_side = Side(style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in worksheet[header_row]:
        if cell.column <= last_col:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, max_col=last_col):
        is_total_row = any(str(cell.value) == "Grand Total" for cell in row[:3])
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if is_total_row:
                cell.fill = total_fill
                cell.font = bold_font
            elif cell.column <= 3:
                cell.fill = metric_fill

    if freeze:
        worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=2)
    for column_index in range(1, last_col + 1):
        max_len = max(
            len(str(worksheet.cell(row=row_index, column=column_index).value or ""))
            for row_index in range(header_row, last_row + 1)
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 11), 18)


def add_table_block(worksheet, title, output_df, start_row, freeze=True):
    worksheet.cell(row=start_row, column=1, value=title)
    worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14, color="1F4E78")
    header_row = start_row + 1

    for row_index, row_values in enumerate(dataframe_to_rows(output_df, index=False, header=True), header_row):
        for column_index, value in enumerate(row_values, 1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    style_excel_range(worksheet, header_row, header_row + len(output_df), len(output_df.columns), freeze=freeze)
    return header_row, header_row + len(output_df), len(output_df.columns)


def add_excel_chart(worksheet, title, chart_kind, header_row, last_row, last_col, anchor):
    headers = [worksheet.cell(row=header_row, column=col).value for col in range(1, last_col + 1)]
    overlay_candidates = ["Target Stock", "Tester Capacity", "Next Month Demand"]
    overlay_col = next((headers.index(col) + 1 for col in overlay_candidates if col in headers), None)
    chart_last_col = overlay_col - 1 if overlay_col else last_col
    if worksheet.cell(row=header_row, column=chart_last_col).value == "Grand Total":
        chart_last_col -= 1
    chart_last_row = last_row
    if worksheet.cell(row=chart_last_row, column=1).value == "Grand Total":
        chart_last_row -= 1
    if chart_last_col < 2 or chart_last_row <= header_row:
        return

    chart_map = {"bar": BarChart, "line": LineChart, "area": AreaChart, "stock": BarChart, "tester": BarChart}
    chart = chart_map[chart_kind]()
    chart.title = title
    chart.height = 10
    chart.width = 20

    data = Reference(worksheet, min_col=2, max_col=chart_last_col, min_row=header_row, max_row=chart_last_row)
    categories = Reference(worksheet, min_col=1, min_row=header_row + 1, max_row=chart_last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    if chart_kind in ["bar", "stock", "tester"]:
        chart.type = "col"
        chart.style = 10
        if chart_kind == "tester":
            chart.grouping = "stacked"
            chart.overlap = 100
    elif chart_kind == "area":
        chart.grouping = "stacked"

    if chart_kind in ["stock", "tester"] and overlay_col:
        line_chart = LineChart()
        line_data = Reference(worksheet, min_col=overlay_col, max_col=overlay_col, min_row=header_row, max_row=chart_last_row)
        line_chart.add_data(line_data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.y_axis.axId = 200
        line_chart.y_axis.title = worksheet.cell(row=header_row, column=overlay_col).value
        chart += line_chart

    if chart_kind == "line" and title == "Reach Level":
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 6
        chart.y_axis.majorUnit = 1

    worksheet.add_chart(chart, anchor)


def style_excel_sheet(worksheet):
    style_excel_range(worksheet, 1, worksheet.max_row, worksheet.max_column)


def make_output_bytes(graph_outputs, wafer_start_table, stage_output_table, tester_usage_table, unconstrained_tester_table=None):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wafer_start_table.to_excel(writer, sheet_name="Wafer_Start", index=False)
        stage_output_table.to_excel(writer, sheet_name="Bump_Testing_DPS", index=False)
        tester_usage_table.to_excel(writer, sheet_name="Tester_Allocation", index=False)
        if unconstrained_tester_table is not None:
            unconstrained_tester_table.to_excel(writer, sheet_name="Tester_Required_No_Limit", index=False)
        style_excel_sheet(writer.book["Wafer_Start"])
        style_excel_sheet(writer.book["Bump_Testing_DPS"])
        style_excel_sheet(writer.book["Tester_Allocation"])
        if unconstrained_tester_table is not None:
            style_excel_sheet(writer.book["Tester_Required_No_Limit"])

        tester_graph_outputs = [item for item in graph_outputs if item[0] == "Tester Used"]
        other_graph_outputs = [item for item in graph_outputs if item[0] != "Tester Used"]

        tester_graph_sheet = writer.book.create_sheet("Tester_Graph", 0)
        for index, (title, graph_df, chart_kind) in enumerate(tester_graph_outputs):
            start_row = 1 + (index * 29)
            header_row, last_row, last_col = add_table_block(tester_graph_sheet, title, graph_df, start_row, freeze=False)
            add_excel_chart(tester_graph_sheet, title, chart_kind, header_row, last_row, last_col, f"L{start_row}")
        tester_graph_sheet.freeze_panes = None

        other_graph_sheet = writer.book.create_sheet("Other_Graphs", 1)
        for index, (title, graph_df, chart_kind) in enumerate(other_graph_outputs):
            start_row = 1 + (index * 29)
            header_row, last_row, last_col = add_table_block(other_graph_sheet, title, graph_df, start_row, freeze=False)
            add_excel_chart(other_graph_sheet, title, chart_kind, header_row, last_row, last_col, f"L{start_row}")
        other_graph_sheet.freeze_panes = None
    output.seek(0)
    return output.getvalue()


def style_display_table(output_df):
    numeric_cols = output_df.select_dtypes(include="number").columns.tolist()
    metric_colors = {
        "Bump": "background-color: #eef7f2",
        "Testing": "background-color: #eef3fb",
        "DPS": "background-color: #fff6e8",
    }

    def highlight_total(row):
        is_total = any(str(value) == "Grand Total" for value in row.values)
        if is_total:
            return ["font-weight: 700; background-color: #f2f4f7" for _ in row]

        metric = str(row.get("Metric", ""))
        row_color = metric_colors.get(metric, "")
        return [row_color for _ in row]

    return (
        output_df.style
        .format({col: "{:,.0f}" for col in numeric_cols})
        .set_table_styles([
            {"selector": "th", "props": [("background-color", "#eef3f8"), ("color", "#17202a"), ("font-weight", "700"), ("border", "1px solid #d0d7de")]},
            {"selector": "td", "props": [("background-color", "#ffffff"), ("border", "1px solid #d0d7de")]},
        ])
        .apply(highlight_total, axis=1)
    )


def draw_graphs(graph_outputs):
    st.subheader("Graph")
    for title, graph_df, chart_kind in graph_outputs:
        chart_df = graph_df[graph_df["Row Labels"].astype(str) != "Grand Total"].copy()
        x_order = chart_df["Row Labels"].astype(str).tolist()

        if chart_kind == "tester":
            product_cols = [col for col in chart_df.columns if col not in ["Row Labels", "Grand Total", "Tester Capacity"]]
            tester_long_df = chart_df.melt(id_vars="Row Labels", value_vars=product_cols, var_name="Product", value_name="Tester Used")
            tester_long_df = tester_long_df[tester_long_df["Tester Used"] > 0]
            chart = px.bar(tester_long_df, x="Row Labels", y="Tester Used", color="Product", title=title)
            chart.update_layout(barmode="stack")
            chart.add_scatter(
                x=chart_df["Row Labels"],
                y=chart_df["Tester Capacity"],
                mode="lines+markers",
                name="Tester Capacity",
                line={"color": "#d62728", "width": 3},
                marker={"size": 8},
            )
            chart.update_layout(xaxis_title="Week", yaxis_title="Tester")
            chart.update_xaxes(type="category", categoryorder="array", categoryarray=x_order)
            st.plotly_chart(chart, width="stretch")
            continue

        if chart_kind == "stock":
            target_line = chart_df[["Row Labels", "Target Stock"]].copy()
            product_cols = [col for col in chart_df.columns if col not in ["Row Labels", "Grand Total", "Target Stock"]]
            chart_df = chart_df.melt(id_vars="Row Labels", value_vars=product_cols, var_name="Product", value_name=title)
            chart = px.bar(chart_df, x="Row Labels", y=title, color="Product", title=title)
            chart.update_layout(barmode="stack")
            chart.add_scatter(
                x=target_line["Row Labels"],
                y=target_line["Target Stock"],
                mode="lines+markers",
                name="Target Stock",
                line={"color": "#d62728", "width": 3},
                marker={"size": 8},
            )
            chart.update_layout(xaxis_title="Month", yaxis_title="Stock / Target Stock")
            chart.update_xaxes(type="category", categoryorder="array", categoryarray=x_order)
            st.plotly_chart(chart, width="stretch")
            continue

        chart_df = chart_df.melt(id_vars="Row Labels", var_name="Product", value_name=title)
        chart_df = chart_df[chart_df["Product"] != "Grand Total"]

        if chart_kind == "line":
            chart = px.line(chart_df, x="Row Labels", y=title, color="Product", markers=True, title=title)
            if title == "Reach Level":
                chart.update_yaxes(range=[0, 6], dtick=1)
        else:
            chart = px.bar(chart_df, x="Row Labels", y=title, color="Product", title=title)
            chart.update_layout(barmode="stack")

        chart.update_layout(xaxis_title="Month", yaxis_title=title)
        chart.update_xaxes(type="category", categoryorder="array", categoryarray=x_order)
        st.plotly_chart(chart, width="stretch")


uploaded = st.sidebar.file_uploader("Upload input Excel", type=["xlsx"])
file_bytes = uploaded.getvalue() if uploaded else None

try:
    tables = load_tables(file_bytes)
except Exception as exc:
    st.error(str(exc))
    st.stop()

if tables is None:
    st.info("Please upload an input Excel file from the sidebar, or add input_sample.xlsx to the repo.")
    st.stop()

tables["product"] = ensure_priority_column(tables["product"])

demand_source_key = uploaded.name if uploaded else "default_input"
if st.session_state.get("simple_demand_source") != demand_source_key:
    st.session_state["simple_demand_source"] = demand_source_key
    st.session_state["simple_demand_df"] = tables["demand"].copy()
    st.session_state["simple_demand_version"] = 0
elif "simple_demand_df" not in st.session_state:
    st.session_state["simple_demand_df"] = tables["demand"].copy()
    st.session_state["simple_demand_version"] = 0

min_reach_default = float(target_value(tables["target"], "Min Reach Level", planner.DEFAULT_TESTER_CONFIG["min_REACH"]))
target_reach_default = float(target_value(tables["target"], "Target Reach Level", planner.DEFAULT_TESTER_CONFIG["target_REACH"]))
max_reach_default = float(target_value(tables["target"], "Max Reach Level", planner.DEFAULT_TESTER_CONFIG["max_REACH"]))
tester_number_default = int(target_value(tables["target"], "Tester Number", planner.DEFAULT_TESTER_CONFIG["available"]))
tester_smoothing_default = int(target_value(tables["target"], "Tester Smoothing Weeks", planner.DEFAULT_TESTER_CONFIG["tester_smoothing_weeks"]))
wafer_start_before_default = target_value(tables["target"], "Wafer Start Time", None)
wafer_start_before_default = "" if pd.isna(wafer_start_before_default) else str(int(wafer_start_before_default))

with st.sidebar:
    st.header("Planning Inputs")
    min_reach = st.number_input("Min Reach Level", min_value=0.0, value=min_reach_default, step=0.5)
    target_reach = st.number_input("Target Reach Level", min_value=0.0, value=target_reach_default, step=0.5)
    max_reach = st.number_input("Max Reach Level", min_value=0.0, value=max_reach_default, step=0.5)
    tester_number = st.number_input("Tester Number", min_value=1, value=tester_number_default, step=1)
    tester_smoothing_weeks = st.number_input(
        "Tester Smoothing Weeks",
        min_value=0,
        value=tester_smoothing_default,
        step=1,
        help="Allows future tester load to be pulled earlier into idle tester weeks. Set to 0 to disable smoothing.",
    )
    wafer_start_before_text = st.text_input(
        "Wafer Start Time",
        value=wafer_start_before_default,
        help="Optional. Enter how many weeks before demand wafer start is allowed. Leave blank for no limit.",
    )

tabs = st.tabs(["Flow", "Product", "Demand", "Inventory"])
with tabs[0]:
    st.caption("Transit time is shown as a column on each stage. When you run the plan, it will be converted back to Transist Time rows automatically.")
    flow_editor_df = flow_to_editor_table(tables["flow"])
    flow_headers = render_header_editor("simple_flow", flow_editor_df.columns, demand_source_key)
    flow_df = st.data_editor(
        flow_editor_df,
        width="stretch",
        num_rows="dynamic",
        key="simple_flow",
        column_config={
            "Stage": st.column_config.TextColumn(flow_headers.get("Stage", "Stage"), help="Main process step, excluding transit rows."),
            "Cycle Time Week": st.column_config.NumberColumn(flow_headers.get("Cycle Time Week", "Cycle Time Week"), min_value=0.0, step=1.0, format="%.0f"),
            "Transit Week": st.column_config.NumberColumn(flow_headers.get("Transit Week", "Transit Week"), min_value=0.0, step=1.0, format="%.0f"),
            "Yield": st.column_config.NumberColumn(flow_headers.get("Yield", "Yield"), min_value=0.0, max_value=1.0, step=0.001, format="%.3f"),
        },
    )
with tabs[1]:
    product_headers = render_header_editor("simple_product", tables["product"].columns, demand_source_key)
    product_column_config = generic_column_config(tables["product"].columns, product_headers)
    if "Priority" in tables["product"].columns:
        product_column_config["Priority"] = st.column_config.NumberColumn(
            product_headers.get("Priority", "Priority"),
            help="Higher number means higher priority. When weekly total tester capacity is not enough, products with higher Priority are allocated first.",
            step=1,
            format="%.0f",
        )
    product_df = st.data_editor(
        tables["product"],
        width="stretch",
        num_rows="dynamic",
        key="simple_product",
        column_config=product_column_config,
    )
with tabs[2]:
    st.caption("Choose the demand date range. Existing month values are kept; new months are added with 0 demand.")
    default_start, default_end = demand_date_range_defaults(st.session_state["simple_demand_df"])
    range_col1, range_col2, range_col3 = st.columns([1, 1, 1])
    with range_col1:
        start_month = st.date_input("Start month", value=default_start, key="simple_demand_start_month")
    with range_col2:
        end_month = st.date_input("End month", value=default_end, key="simple_demand_end_month")
    with range_col3:
        st.write("")
        st.write("")
        apply_range = st.button("Apply date range", key="apply_demand_date_range")

    if apply_range:
        try:
            st.session_state["simple_demand_df"] = apply_demand_date_range(
                st.session_state["simple_demand_df"],
                start_month,
                end_month,
            )
            st.session_state["simple_demand_version"] += 1
            st.success("Demand date range updated")
        except Exception as exc:
            st.error(str(exc))

    demand_df = st.data_editor(
        st.session_state["simple_demand_df"],
        width="stretch",
        num_rows="dynamic",
        key=f"simple_demand_{st.session_state['simple_demand_version']}",
        column_config=generic_column_config(
            st.session_state["simple_demand_df"].columns,
            render_header_editor(
                "simple_demand",
                st.session_state["simple_demand_df"].columns,
                f"{demand_source_key}_{st.session_state['simple_demand_version']}",
            )
        ),
    )
    st.session_state["simple_demand_df"] = demand_df.copy()
with tabs[3]:
    inventory_headers = render_header_editor("simple_inventory", tables["inventory"].columns, demand_source_key)
    inventory_df = st.data_editor(
        tables["inventory"],
        width="stretch",
        num_rows="dynamic",
        key="simple_inventory",
        column_config=generic_column_config(tables["inventory"].columns, inventory_headers),
    )

try:
    edited_wafer_start_before = parse_wafer_start_before(wafer_start_before_text)
    edited_input_bytes = make_input_bytes(
        flow_df,
        product_df,
        demand_df,
        inventory_df,
        min_reach,
        target_reach,
        max_reach,
        tester_number,
        tester_smoothing_weeks,
        edited_wafer_start_before,
    )
    st.download_button(
        "Download Edited Input Excel",
        data=edited_input_bytes,
        file_name="edited_input.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as exc:
    st.warning(f"Cannot prepare edited input download: {exc}")

if st.button("Run simple plan", type="primary"):
    with st.spinner("Running simple reach-based plan..."):
        try:
            wafer_start_before = parse_wafer_start_before(wafer_start_before_text)
            if min_reach > target_reach or target_reach > max_reach:
                raise ValueError("Reach levels must satisfy: Min Reach <= Target Reach <= Max Reach.")
            input_path = build_input_workbook(flow_df, product_df, demand_df, inventory_df, min_reach, target_reach, max_reach, tester_number, tester_smoothing_weeks, wafer_start_before)
            stages, products, weekly_demand_map, month_label_map, tester_config, inventory_map = planner.load_all_inputs(input_path)
            results = planner.run_all_products(stages, products, weekly_demand_map, month_label_map, tester_config, inventory_map)
        except Exception as exc:
            st.error(str(exc))
            st.stop()

    all_plan_df, summary_df, monthly_summary_df, skipped_df = results
    graph_outputs = build_graph_outputs(monthly_summary_df, all_plan_df, target_reach, tester_number)
    tester_usage_table = build_tester_usage_table(all_plan_df, tester_number)
    wafer_start_table = build_wafer_start_table(monthly_summary_df)
    stage_output_table = build_stage_output_table(monthly_summary_df)
    output_bytes = make_output_bytes(graph_outputs, wafer_start_table, stage_output_table, tester_usage_table)

    st.success("Plan generated")

    draw_graphs(graph_outputs)

    st.subheader("Wafer Start")
    st.dataframe(style_display_table(wafer_start_table), width="stretch")

    st.subheader("Bump Testing DPS")
    st.dataframe(style_display_table(stage_output_table), width="stretch")

    with st.expander("Tester Allocation"):
        st.dataframe(tester_usage_table, width="stretch")

    with st.expander("Summary"):
        st.dataframe(summary_df, width="stretch")

    if not skipped_df.empty:
        with st.expander("Skipped Products"):
            st.dataframe(skipped_df, width="stretch")

    st.download_button(
        "Download Simple Output Excel",
        data=output_bytes,
        file_name="output_simple_streamlit_result.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
